import openpyxl
import schedule
import time
import xlwings as xw
''''
打开后不支持 ，将文件设为只读试图，也不行
'''
def write_to_excel():
    # 打开现有的Excel文件（如果不存在则创建新文件）
    try:
        workbook = xw.Book('data.xlsx') #openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        # workbook = openpyxl.Workbook()
        workbook = xw.Book()

    # 选择默认的工作表（如果不存在则创建新工作表）
    # sheet = workbook.active
    # 获取特定的Sheet
    sheet = workbook.sheets['Sheet1']  # 修改为你的Sheet名称

    # 获取当前时间
    current_time = time.strftime("%Y-%m-%d %H:%M:%S")

    # 在下一空行中写入数据
    # next_row = sheet.max_row + 1
    sheet.cell(row=1, column=1, value=current_time)
    sheet.cell(row=2, column=2, value="New Data")
    sheet.cell(row=3, column=3, value="New Data3")

    # 保存Excel文件
    workbook.save('data.xlsx')

    # 显式地关闭Excel文件
    workbook.close()

# 每隔一分钟执行一次写入操作
schedule.every(5).seconds.do(write_to_excel)

# 无限循环以便定时任务持续执行
while True:
    schedule.run_pending()
    time.sleep(1)
