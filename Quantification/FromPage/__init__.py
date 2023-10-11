import requests
from openpyxl.reader.excel import load_workbook

'''
自己尝试抓，不熟练
'''
# url = 'http://www.iwencai.com/stockpick/load-data?rsh=3&typed=0&preParams=&ts=1&f=1&qs=result_original&selfsectsn=&querytype=stock&searchfilter=&tid=stockpick&w=%E4%BB%8A%E6%97%A5%E6%B6%A8%E5%81%9C%E5%92%8C%E6%B6%A8%E5%81%9C%E5%8E%9F%E5%9B%A0%E9%9D%9EST+%E9%A6%96%E6%AC%A1%E6%B6%A8%E5%81%9C%E6%97%B6%E9%97%B4%E6%8E%92%E5%BA%8F&queryarea='
# response = requests.get(url)
#
# # 打印网页内容
# print(response.text)
'''
借助问财 
https://github.com/zsrl/pywencai

调试关键词，让其匹配同花顺对应查询条件  -- prompt
'行业简称', --不确定匹配上
定时任务 +excel
'''

import pywencai
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw
def get():




    # print("=========== zt =============")
    current_date = date.today()
    cur_date_str =  '[' + str(current_date.strftime("%Y%m%d")) + ']'

    res_zt = pywencai.get(query='昨日涨停和涨停原因非ST 首次涨停时间排序 行业', sort_key='', sort_order='涨停类型')
    filter_col_zt = ['code','股票简称','最新价',
                  '首次涨停时间'+cur_date_str,'涨停原因类别'+cur_date_str,'涨停类型'+cur_date_str,
                  '连续涨停天数'+cur_date_str,'几天几板' + cur_date_str,
                  '涨停封单量'+cur_date_str,'涨停封单额'+cur_date_str,'涨停开板次数'+cur_date_str

                  ]
    # print(res[filter_col])
    # file_name = 'zt.csv'
    # res_zt[filter_col].to_csv('./data/' + file_name, index=False)


    # print("=========== >%5 =============")
    res_dy5 = pywencai.get(query='昨日涨幅大于5 非30开头 非 688开头 非ST 非涨停 展示行业 ', sort_key='', sort_order='')
    # print(res)
    filter_col_dy5 = ['code','股票简称','最新价',
                  '成交量' + cur_date_str, '所属同花顺行业',
                  '涨停价'+ cur_date_str
                  ]
    # file_name = 'dy5.csv'
    # res[filter_col].to_csv('./data/' + file_name, index=False)



    # print("=========== zb =============")
    res_zb = pywencai.get(query='昨日炸板不含ST 所属行业', sort_key='', sort_order='')
    filter_col_zb = ['code','股票简称','最新价',
                  '涨停时间明细'+cur_date_str,'涨停开板次数'+cur_date_str,
                  '所属同花顺行业',
                  '涨停价'+ cur_date_str
                  ]

    # file_name = 'zb.csv'
    # res[filter_col].to_csv('./data/' + file_name, index=False)
    # print(res)

    a_hx = ['000010', '600519']
    ret_hx = pywencai.get(query='当前人气排名 行业 ', find=a_hx)
    filter_col_hx = ['code', '股票简称', '个股热度排名'+cur_date_str ,'最新涨跌幅','最新价',
                    '所属同花顺行业','所属概念'
                     ]

    #写到一个excel不同sheet
    # with pd.ExcelWriter('./data/stronger.xlsx') as writer:
    #     res_zt[filter_col_zt].to_excel(writer, sheet_name='Sheet1', index=False)
    #     res_dy5[filter_col_dy5].to_excel(writer, sheet_name='Sheet2', index=False)
    #     res_zb[filter_col_zb].to_excel(writer, sheet_name='Sheet3', index=False)
    #     ret_hx.to_excel(writer, sheet_name='Sheet4', index=False)

    # 创建一个新的工作簿
    # workbook = Workbook()
    #
    # # 获取当前活动的工作表
    # sheet = workbook.active
    #
    # # # 向单元格写入数据
    # sheet['A1'] = 'Hello'
    # sheet['B1'] = 'World'
    #
    '''
    直接load可以不 - 不行 --不行
    '''
    # 获取当前活动的工作表
    # first_sheet = workbook.active

    # # 保存工作簿
    # workbook.save('./data/stronger.xlsx')
    # 创建一个Excel工作簿
    # workbook = Workbook()
    try:
        #每次都打开新的，相当于半个创建
        workbook = xw.Book('./data/stronger.xlsx')

        # 获取已经运行的Excel应用程序实例，如果没有运行的实例，则会启动一个新的Excel应用程序
        # app = xw.apps.active
        # 打开现有的Excel工作簿，如果文件不存在，则会创建一个新的工作簿
        # workbook = app.books.open('./data/stronger.xlsx')  #修改为你的Excel文件路径

        print("--1--")
    except FileNotFoundError:
        print("--2--")
        workbook = xw.Book()
        # 创建其他工作表并指定名称
        sheet_names = ['Sheet2', 'Sheet3', 'Sheet4','Sheet5']
        # 使用循环创建并指定多个工作表的名称
        for sheet_name in sheet_names:
            workbook.sheets.add(sheet_name,after=workbook.sheets[-1])
            # workbook.create_sheet(title=sheet_name)
        # 必须保存，否则加载找不到
        workbook.save('./data/stronger.xlsx')
    # sheet = workbook.active

    # 创建WPS表格应用程序实例
    # app = xw.App(visible=True, add_book=False, spec='wps')

    # 打开现有的Excel工作簿，如果文件不存在，则会创建一个新的工作簿
    # try:
    #     workbook = app.books.open('./data/stronger.xlsx')  # 修改为你的Excel文件路径
    # except Exception as e:
    #     print(f"Error: {e}")
    # sheet_names = ['Sheet2', 'Sheet3', 'Sheet4','Sheet5']
    # # 使用循环创建并指定多个工作表的名称
    # for sheet_name in sheet_names:
    #     # workbook.sheets.add(sheet_name,after=workbook.sheets[-1])
    #     # workbook.create_sheet(title=sheet_name)
    #     new_sheet = workbook.sheets.add(sheet_name, after=workbook._sheets[-1])

        # 创建第一个sheet页并写入第一个DataFrame数据
    # sheet1 = workbook.active
    # sheet1.title = 'Sheet1'

    # 获取工作簿中的所有工作表
    # all_sheets = workbook._sheets
    # print(all_sheets)

    sheet1=workbook.sheets['Sheet2']          #.sheets[0]
    # for row in dataframe_to_rows(res_zt, index=False, header=True):#[filter_col_zt]
    #     # sheet1.append(row)
    #     sheet1.range.value

        # # 在空白行写入数据
        # for i, data in enumerate(row_data):
        #     sheet1.range(f'A{empty_row + i}').value = data

    # 将DataFrame写入Excel的指定单元格
    # workbook.get_sheet_by_name("Sheet1").
    sheet1.range('A1').value = res_zt  # 将DataFrame写入A1单元格，xlwings会自动处理整个DataFrame
    #
    # # 创建第二个sheet页并写入第二个DataFrame数据
    # # sheet2 = workbook.create_sheet(title='Sheet2')
    # sheet2 = workbook.active
    # sheet2.title = 'Sheet2'
    sheet2=workbook.sheets['Sheet3']
    # for row in dataframe_to_rows(res_dy5, index=False, header=True):#[filter_col_dy5]
    #     sheet2.append(row)
    sheet2.range('A1').value = res_dy5
    #
    # sheet3 = workbook.active
    # sheet3.title = 'Sheet3'
    sheet3=workbook.sheets['Sheet4']
    # for row in dataframe_to_rows(res_zb, index=False, header=True):#[filter_col_zb]
    #     sheet3.append(row)
    sheet3.range('A1').value = res_zb
    # 创建第二个sheet页并写入第二个DataFrame数据
    # sheet4 = workbook.active
    # sheet4.title = 'Sheet4'
    sheet4=workbook.sheets['Sheet5']
    # for row in dataframe_to_rows(ret_hx, index=False, header=True):#[filter_col_hx]
    #     sheet4.append(row)
    sheet4.range('A1').value = ret_hx
    # 保存Excel文件
    print('fake save --')
    # workbook.save('./data/stronger.xlsx')    save会导致出现每次一个文件

    # 显式地关闭Excel文件
    # workbook.close()


